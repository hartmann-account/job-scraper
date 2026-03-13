#!/usr/bin/env python3
"""
Job Scraper — Monatliche Automatisierung
=========================================
Orchestriert alle Scraper-Module, dedupliziert Ergebnisse,
und speichert sie als CSV (oder XLSX).

Ausführung:
    python main.py                  # Normaler Lauf
    python main.py --dry-run        # Testlauf ohne Speichern
    python main.py --format xlsx    # Output als Excel
"""

import argparse
import csv
import logging
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Dict

import yaml

# Scraper-Module
from scrapers import arbeit_swiss, linkedin, career_pages

# ─── Logging ────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("scraper.log", encoding="utf-8")
    ]
)
logger = logging.getLogger(__name__)

# ─── Konstanten ─────────────────────────────────────────────
OUTPUT_DIR = Path("output")
FIELDNAMES = [
    "source", "title", "company", "location",
    "workload", "published", "url", "scraped_at"
]


def load_config(path: str = "config.yaml") -> dict:
    """Lädt die YAML-Konfiguration."""
    config_path = Path(path)
    if not config_path.exists():
        logger.error(f"Config-Datei nicht gefunden: {path}")
        sys.exit(1)

    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)

    logger.info(f"Konfiguration geladen: {len(config.get('search_terms', []))} Suchbegriffe, "
                f"{len(config.get('locations', []))} Standorte")
    return config


def deduplicate(results: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """Entfernt Duplikate basierend auf URL und Titel+Firma."""
    seen = set()
    unique = []

    for job in results:
        # Dedupliziere über URL (primär) und Titel+Firma (sekundär)
        key_url = job.get("url", "").lower().strip()
        key_title = f"{job.get('title', '').lower()}|{job.get('company', '').lower()}"

        if key_url != "n/a" and key_url in seen:
            continue
        if key_title in seen:
            continue

        seen.add(key_url)
        seen.add(key_title)
        unique.append(job)

    return unique


def save_csv(results: List[Dict[str, str]], filepath: Path):
    """Speichert Ergebnisse als CSV."""
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES, delimiter=";")
        writer.writeheader()
        writer.writerows(results)
    logger.info(f"CSV gespeichert: {filepath} ({len(results)} Einträge)")


def save_xlsx(results: List[Dict[str, str]], filepath: Path):
    """Speichert Ergebnisse als XLSX (braucht openpyxl)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl nicht installiert. Fallback auf CSV.")
        save_csv(results, filepath.with_suffix(".csv"))
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    # Header
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    for col_idx, field in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col_idx, value=field.replace("_", " ").title())
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Daten
    for row_idx, job in enumerate(results, 2):
        for col_idx, field in enumerate(FIELDNAMES, 1):
            ws.cell(row=row_idx, column=col_idx, value=job.get(field, ""))

    # Spaltenbreiten
    col_widths = [15, 45, 25, 20, 10, 15, 60, 22]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Autofilter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    logger.info(f"XLSX gespeichert: {filepath} ({len(results)} Einträge)")


def main():
    parser = argparse.ArgumentParser(description="Monatlicher Job-Scraper")
    parser.add_argument("--config", default="config.yaml", help="Pfad zur Konfiguration")
    parser.add_argument("--format", choices=["csv", "xlsx"], default=None, help="Output-Format")
    parser.add_argument("--dry-run", action="store_true", help="Testlauf ohne Speichern")
    args = parser.parse_args()

    logger.info("=" * 60)
    logger.info("JOB SCRAPER — Monatlicher Lauf")
    logger.info(f"Zeitpunkt: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 60)

    config = load_config(args.config)

    # Output-Format aus Config oder CLI
    output_format = args.format or config.get("output", {}).get("format", "csv")

    # ─── Scraper ausführen ──────────────────────────────────
    all_results = []

    logger.info("\n--- Arbeit.swiss ---")
    all_results.extend(arbeit_swiss.scrape(config))

    logger.info("\n--- LinkedIn ---")
    all_results.extend(linkedin.scrape(config))

    logger.info("\n--- Karriereseiten ---")
    all_results.extend(career_pages.scrape(config))

    # ─── Deduplizierung ────────────────────────────────────
    before = len(all_results)
    all_results = deduplicate(all_results)
    after = len(all_results)
    logger.info(f"\nDeduplizierung: {before} → {after} Einträge ({before - after} Duplikate entfernt)")

    # ─── Sortierung: Neueste zuerst ────────────────────────
    all_results.sort(key=lambda x: x.get("published", ""), reverse=True)

    # ─── Speichern ─────────────────────────────────────────
    if args.dry_run:
        logger.info(f"\n[DRY RUN] Würde {len(all_results)} Jobs speichern.")
        for job in all_results[:10]:
            logger.info(f"  {job['source']}: {job['title']} @ {job['company']}")
        if len(all_results) > 10:
            logger.info(f"  ... und {len(all_results) - 10} weitere")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    prefix = config.get("output", {}).get("filename_prefix", "job_scrape")
    date_str = datetime.now().strftime("%Y-%m")
    filename = f"{prefix}_{date_str}.{output_format}"
    filepath = OUTPUT_DIR / filename

    if output_format == "xlsx":
        save_xlsx(all_results, filepath)
    else:
        save_csv(all_results, filepath)

    # ─── Zusammenfassung ───────────────────────────────────
    logger.info("\n" + "=" * 60)
    logger.info("ZUSAMMENFASSUNG")
    logger.info(f"  Ergebnisse: {len(all_results)}")
    logger.info(f"  Output:     {filepath}")

    # Aufschlüsselung nach Quelle
    sources = {}
    for job in all_results:
        src = job["source"]
        sources[src] = sources.get(src, 0) + 1
    for src, count in sorted(sources.items()):
        logger.info(f"  {src}: {count}")

    logger.info("=" * 60)


if __name__ == "__main__":
    main()
